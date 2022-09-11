
import os
import csv
import sys
import json
import time
import atexit
import argparse
import tabulate
import pyperclip

import tkinter as tk

from pathlib import Path
from hashlib import sha256
from pynput import keyboard
from openpyxl import load_workbook
from screeninfo import get_monitors
from win10toast import ToastNotifier
from tkinter import filedialog, simpledialog, messagebox

parser = argparse.ArgumentParser(description="Optional flags")
parser.add_argument("--force-notify", help="Display toast notifications on multiple monitors.", action=argparse.BooleanOptionalAction, type=bool, default=False)
parser.add_argument("--disable-toast", help="Disable toast notifications.", action=argparse.BooleanOptionalAction, type=bool, default=False)
parser.add_argument("--disable-popup", help="Disable popup warnings.", action=argparse.BooleanOptionalAction, type=bool, default=False)
args = parser.parse_args()

def showerr(title, message):
    if not vars(args)['disable_popup']:
        messagebox.showerror(title, message)

def showinf(title, message):
    if not vars(args)['disable_popup']:
        messagebox.showinfo(title, message)

def showwarn(title, message):
    if not vars(args)['disable_popup']:
        messagebox.showwarning(title, message)

def CSVtoJSON(path):
    data = []
    with open(path,"r") as f:
        csvReader = csv.DictReader(f)
        data.append({x:x for x in csvReader.fieldnames})
        for rows in csvReader:
            data.append(rows)
    return data

def read_excel_sheet(abspath, sheet_name):
    workbook = load_workbook(abspath, data_only=True)
    worksheet = workbook[sheet_name]
    return worksheet

def sheet_to_json(worksheet):
    dict_list = []
    header = [cell[0].value for cell in worksheet.columns]
    for row in worksheet.iter_rows():
        row_dict = {col: row[col_idx].internal_value for col_idx, col in enumerate(header)}
        dict_list.append(row_dict)

    return dict_list

def XLSXtoJSON(path,name):
    worksheet = read_excel_sheet(path,name)
    return sheet_to_json(worksheet)


if len(get_monitors()) == 1:
    display_toast = True if not vars(args)("disable_toast") else False
else:
    display_toast = False if not vars(args).get('force_notify') else True

path = Path(__file__).parent

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(title="Select csv file", filetypes=(("CSV files", "*.csv *.txt"), ("Excel files", "*.xlsx"), ("all files", "*.*")))
worksheet_name = ""
if file_path.endswith(".xlsx"):
    worksheet_name = simpledialog.askstring(title="Select worksheet name", prompt="Enter worksheet name")

metadata_path = filedialog.askopenfilename(title="Select metadata file", filetypes=(("JSON files", "*.json"), ("all files", "*.*")))

if file_path == "" or file_path is None:
    showerr(title="Error", message="You must select a file!")
    sys.exit(1)
if (worksheet_name == "" or worksheet_name is None) and file_path.endswith(".xlsx"):
    showerr(title="Error", message="You must select a worksheet name!")
    sys.exit(1)
if file_path.endswith(".xlsx"):
    excel_json = XLSXtoJSON(file_path,worksheet_name)
else:
    excel_json = CSVtoJSON(file_path)

if metadata_path == "" or metadata_path is None:
    showinf(title="No metadata", message="No metadata file selected, using default metadata!")
    with open("metadata.json", "w+") as f:
        f.write("{}")
    h = sha256()
    h.update(json.dumps(excel_json).encode('utf-8'))
    metadata = {h.hexdigest():[0]}
    metadata_path = f"{path}\\metadata.json"
    showinf(title="Success", message="Successfully created metadata file!")
else:
    with open(metadata_path, "r") as f:
        metadata = json.load(f)
h = sha256()
h.update(json.dumps(excel_json).encode("utf-8"))
current_metadata_key = h.hexdigest()
if current_metadata_key in metadata:
    current_key = metadata[current_metadata_key][-1] + 1
else:
    metadata[current_metadata_key] = [0]
    current_key = metadata[current_metadata_key][-1] + 1

showwarn(title="Warning", message="Press * to exit, press N to go to next system, B to go back!")
if display_toast:
    showinf(title="Error", message="Toast notifications are enabled!")
else:
    showerr(title="Error", message="Toast notifications are not supported on multiple monitors!")

original_content = pyperclip.paste()

def safe_exit():
    with open(metadata_path, "w+") as f:
        json.dump(metadata, f)
    pyperclip.copy(original_content)

atexit.register(safe_exit)

def on_press(key):
    try: 
        if key.char == "*":
            sys.exit()
        elif key.char == "n":
            render_next()
        elif key.char == "b":
            if len(metadata[current_metadata_key]) <= 2:
                return
            metadata[current_metadata_key].pop()
            metadata[current_metadata_key].pop()
            render_next("Displaying previous!")
    except AttributeError:
        pass

toast = ToastNotifier()

def render_next(msg="Displaying next!"):
    os.system("cls")
    if len(metadata[current_metadata_key])+1 == len(excel_json):
        print("You are done!")
        return
    print(msg)
    current_key = metadata[current_metadata_key][-1] + 1
    metadata[current_metadata_key].append(len(metadata[current_metadata_key]))
    temp = json.loads(json.dumps(excel_json[current_key]))
    temp["Body Name"] = temp["Body Name"].partition(temp["System Name"])[2]
    if temp["System Name"] == excel_json[current_key - 1]["System Name"] and msg == "Displaying next!":
        temp["System Name"] = "Current"
        if display_toast:
            try:
                toast.show_toast("Current system", f"Body: {temp['Body Name']}", duration=5,threaded=True)
            except:
                time.sleep(5)
                toast.show_toast("Current system", f"Body: {temp['Body Name']}", duration=5,threaded=True)
    else:
        pyperclip.copy(temp["System Name"])
        if display_toast:
            try:
                toast.show_toast(f"Head to {temp['System Name']}", f"Body: {temp['Body Name']}", duration=5,threaded=True)
            except:
                time.sleep(5)
                toast.show_toast(f"Head to {temp['System Name']}", f"Body: {temp['Body Name']}", duration=5,threaded=True)
    del temp["Distance To Arrival"]
    del temp["Jumps"]
    print(temp["Estimated Scan Value"])
    temp["Estimated Scan Value"] = str('{:,}'.format(int(temp["Estimated Scan Value"])).replace(',', ' '))
    temp["Estimated Mapping Value"] = str('{:,}'.format(int(temp["Estimated Mapping Value"])).replace(',', ' '))
    table = [
        ["System Name", "Body Name", "Body Subtype", "Is Terraformable", "Estimated Scan Value", "Estimated Mapping Value"],
        temp.values()
    ]
    print(tabulate.tabulate(table,headers="firstrow",tablefmt="pretty"))
    print(f'{current_key}/{len(excel_json)} | {str(f"{current_key/len(excel_json)*100:.2f}")}%')
    s = 0
    for i in range(1, current_key):
        s += int(excel_json[i]["Estimated Scan Value"]) + int(excel_json[i]["Estimated Mapping Value"])
    print(f"Total estimated value: {'{:,}'.format(s).replace(',', ' ')}")
    current_key += 1
    
render_next()
    
with keyboard.Listener(on_press=on_press) as listener:
    listener.join() #riches-Chona-Brestla-AE3F31 (2)
